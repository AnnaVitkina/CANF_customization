import json
import math
import pandas as pd
import openpyxl
import os


def process_rate_card(file_path):
    """
    Process a Rate Card Excel file from the input folder.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder (e.g., "rate_card.xlsx")
    
    Returns:
        tuple: (dataframe, list of column names, conditions dictionary)
            - dataframe: Processed pandas DataFrame (filtered to black font columns)
            - list: List of column names in the processed dataframe
            - dict: Dictionary of conditions where keys are column names and values are condition text
    """
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Read the Excel file
    df_rate_card = pd.read_excel(full_path, sheet_name="Rate card", skiprows=2)
    
    # Find first column index (where data actually starts)
    first_column_index = None
    if df_rate_card is not None:
        for i, col in enumerate(df_rate_card.columns):
            if "nan" not in str(df_rate_card.iloc[0, i]).lower():
                first_column_index = i
                break
    
    if first_column_index is not None:
        df_rate_card = df_rate_card.iloc[:, :first_column_index]
    
    # Drop rows where the first column is NaN
    if df_rate_card is not None:
        df_rate_card.dropna(subset=[df_rate_card.columns[0]], inplace=True)
    
    # Set column names from first row
    new_columns = df_rate_card.iloc[0].tolist()
    df_rate_card.columns = new_columns
    df_rate_card = df_rate_card.iloc[1:]
    
    # Load the workbook to extract conditions and check font colors
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    sheet = workbook["Rate card"]
    
    # Find the header row that contains "Currency"
    first_data_row_index = None
    currency_index = None
    
    for row_index in range(1, min(10, sheet.max_row + 1)):
        row = sheet[row_index]
        row_values = [cell.value for cell in row]
        if "Currency" in row_values:
            currency_index = row_values.index("Currency")
            first_data_row_index = row_index
            break
    
    black_font_values = []
    column_notes = {}  # Will store conditions/notes for each column
    
    if first_data_row_index is not None and currency_index is not None:
        # Access the data in this row
        first_data_row = sheet[first_data_row_index]
        first_data_values = [cell.value for cell in first_data_row]
        truncated_data_values = first_data_values[:currency_index]
        
        # Extract conditional rules/notes from COMMENTS in the header row cells
        header_row_index = first_data_row_index
        if header_row_index and header_row_index <= sheet.max_row:
            for i, col_name in enumerate(truncated_data_values, 1):
                if col_name:  # Only process non-empty column names
                    header_cell = sheet.cell(row=header_row_index, column=i)
                    
                    # Check for comments (where conditional rules are stored)
                    if header_cell.comment:
                        comment_text = header_cell.comment.text
                        if comment_text and comment_text.strip():
                            column_notes[col_name] = comment_text.strip()
                    
                    # Also check for cell value notes (in case some are stored as values in row 2)
                    if col_name not in column_notes:
                        notes_row_index = 2
                        if notes_row_index <= sheet.max_row:
                            note_cell = sheet.cell(row=notes_row_index, column=i)
                            if note_cell.value and str(note_cell.value).strip():
                                column_notes[col_name] = str(note_cell.value).strip()
        
        # Check font color to identify black font columns (required columns)
        # Track column indices with black font to handle duplicate column names
        black_font_indices = []  # Store (index, column_name) tuples
        
        for i, value in enumerate(truncated_data_values):
            if i < len(first_data_row):
                cell = first_data_row[i]
                font_color = "black"
                if cell.font and cell.font.color:
                    hex_color = cell.font.color.rgb
                    if hex_color is not None:
                        # Convert to string and handle different formats
                        hex_str = str(hex_color).upper()
                        # Remove 'FF' prefix if present (ARGB format)
                        if hex_str.startswith('FF') and len(hex_str) == 8:
                            hex_str = hex_str[2:]
                        
                        # Check if it's black
                        if hex_str == '000000' or hex_str == '00000000':
                            font_color = "black"
                        else:
                            # Check if it's a shade of grey (R, G, and B are close)
                            try:
                                if len(hex_str) == 6:
                                    r, g, b = int(hex_str[0:2], 16), int(hex_str[2:4], 16), int(hex_str[4:6], 16)
                                    # Check if it's a shade of grey (R, G, and B are close)
                                    if abs(r - g) < 10 and abs(g - b) < 10 and r > 0:  # Grey (not black, not white)
                                        font_color = "grey"
                                    else:
                                        font_color = "other non-black"  # For colors that are not black or grey
                            except (ValueError, IndexError):
                                pass
                
                if font_color == "black":
                    black_font_values.append(value)
                    black_font_indices.append(i)  # Track the index
    
    # Filter the DataFrame to keep only the columns with black font
    # Handle duplicate column names by using positional selection
    if df_rate_card is not None and black_font_indices:
        # Get the original column positions in the dataframe
        # The dataframe columns should correspond to truncated_data_values after setting columns from row
        
        # Check for duplicate column names and keep only black font versions
        seen_columns = {}  # {column_name: index_in_black_font_indices}
        indices_to_keep = []
        
        for idx in black_font_indices:
            col_name = truncated_data_values[idx] if idx < len(truncated_data_values) else None
            if col_name is not None:
                if col_name not in seen_columns:
                    # First occurrence of this column name with black font
                    seen_columns[col_name] = idx
                    indices_to_keep.append(idx)
                else:
                    # Duplicate column name - skip it (keep only the first black font occurrence)
                    print(f"   [DEBUG] Duplicate column '{col_name}' found at index {idx}, keeping first occurrence at index {seen_columns[col_name]}")
        
        # Select columns by position using iloc
        if indices_to_keep:
            # Map the indices to dataframe column positions
            # The dataframe was truncated to first_column_index columns
            valid_indices = [i for i in indices_to_keep if i < len(df_rate_card.columns)]
            if valid_indices:
                df_filtered_rate_card = df_rate_card.iloc[:, valid_indices]
                # Update black_font_values to match the filtered columns
                black_font_values = [truncated_data_values[i] for i in valid_indices if i < len(truncated_data_values)]
            else:
                df_filtered_rate_card = df_rate_card
        else:
            df_filtered_rate_card = df_rate_card
    else:
        df_filtered_rate_card = df_rate_card
    
    # Get list of column names
    column_names = df_filtered_rate_card.columns.tolist()
    
    # Create conditions dictionary (only for columns that exist in the filtered dataframe)
    # Use cleaned conditions for better matching
    conditions = {}
    for col_name in column_names:
        if col_name in column_notes:
            # Clean the condition text for better parsing
            raw_condition = column_notes[col_name]
            cleaned_condition = clean_condition_text(raw_condition)
            conditions[col_name] = cleaned_condition
            
            # Debug: Print raw vs cleaned conditions
            print(f"   [DEBUG] Condition for '{col_name}':")
            print(f"      Raw: {raw_condition[:80]}..." if len(raw_condition) > 80 else f"      Raw: {raw_condition}")
            print(f"      Cleaned: {cleaned_condition[:80]}..." if len(cleaned_condition) > 80 else f"      Cleaned: {cleaned_condition}")
    
    return df_filtered_rate_card, column_names, conditions


def process_business_rules(file_path):
    """
    Process the Business rules tab from a Rate Card Excel file.
    
    Args:
        file_path (str): Path to the file relative to the "input/" folder
    
    Returns:
        dict: Dictionary containing:
            - 'postal_code_zones': list of zone rules with name, country, postal_codes, exclude
            - 'country_regions': list of region rules with name, country, postal_codes, exclude
            - 'no_data_added': list of entries with no data
            - 'raw_rules': all parsed rules as a list of dicts
    """
    import re
    
    # Construct full path from input folder
    input_folder = "input"
    full_path = os.path.join(input_folder, file_path)
    
    # Load the workbook
    workbook = openpyxl.load_workbook(full_path, data_only=True)
    
    # Check if "Business rules" sheet exists
    if "Business rules" not in workbook.sheetnames:
        print(f"   [WARNING] 'Business rules' sheet not found in {file_path}")
        return {
            'postal_code_zones': [],
            'country_regions': [],
            'no_data_added': [],
            'raw_rules': []
        }
    
    sheet = workbook["Business rules"]
    
    # DEBUG: Print sheet info
    print(f"\n{'='*60}")
    print(f"[DEBUG] BUSINESS RULES SHEET ANALYSIS")
    print(f"{'='*60}")
    print(f"   Sheet name: 'Business rules'")
    print(f"   Total rows in sheet: {sheet.max_row}")
    print(f"   Total columns: {sheet.max_column}")
    print(f"   Available sheets: {workbook.sheetnames}")
    
    # STEP 1: Read all rows and filter out empty ones (skip first 2 rows)
    print(f"\n   [DEBUG] Step 1: Reading and filtering rows (skipping first 2 rows)...")
    
    all_rows = []  # Will store (original_row_idx, row_values) tuples
    for row_idx in range(3, sheet.max_row + 1):
        row = sheet[row_idx]
        row_values = [cell.value for cell in row]
        
        # Check if row is empty
        is_empty = all(v is None or (isinstance(v, str) and v.strip() == '') for v in row_values)
        
        if not is_empty:
            all_rows.append((row_idx, row_values))
    
    print(f"   [DEBUG] Total non-empty rows found: {len(all_rows)} (out of {sheet.max_row - 2} after skipping first 2)")
    
    # DEBUG: Print first 20 non-empty rows to see structure
    print(f"\n   [DEBUG] First 20 non-empty rows content:")
    for i, (row_idx, row_values) in enumerate(all_rows[:20]):
        non_empty = [(col_i, v) for col_i, v in enumerate(row_values) if v is not None]
        print(f"      Row {row_idx}: {non_empty}")
    
    if len(all_rows) > 20:
        print(f"      ... and {len(all_rows) - 20} more rows")
    
    # Marker values to look for (case-insensitive)
    markers = ['postal code zones', 'country regions', 'no data added']
    
    # Result structure
    result = {
        'postal_code_zones': [],
        'country_regions': [],
        'no_data_added': [],
        'raw_rules': []
    }
    
    # Track sections and their header columns
    current_section = None
    header_columns = {}  # Maps column index to header name
    waiting_for_header = False  # Flag to indicate we found a marker and are waiting for header row
    
    print(f"\n   [DEBUG] Step 2: Searching for markers: {markers}")
    print(f"   [DEBUG] Structure: MARKER row -> HEADER row (below) -> DATA rows")
    
    # Process non-empty rows
    for i, (row_idx, row_values) in enumerate(all_rows):
        # Check if this row contains a marker (section header)
        row_text_lower = ' '.join(str(v).lower() for v in row_values if v is not None)
        
        found_marker = None
        for marker in markers:
            if marker in row_text_lower:
                found_marker = marker
                print(f"\n   [DEBUG] >>> MARKER FOUND: '{marker}' at row {row_idx}")
                break
        
        if found_marker:
            # This is a marker row - next non-empty row will be the header
            current_section = found_marker.replace(' ', '_')
            waiting_for_header = True
            header_columns = {}  # Reset header columns for new section
            print(f"   [DEBUG]     Section: '{current_section}'")
            print(f"   [DEBUG]     Waiting for header row...")
            continue
        
        # If we're waiting for header, this row should be the header
        if waiting_for_header:
            waiting_for_header = False
            header_columns = {}
            
            print(f"   [DEBUG]     Header row (row {row_idx}): {[v for v in row_values if v is not None]}")
            
            for col_idx, cell_value in enumerate(row_values):
                if cell_value:
                    header_name = str(cell_value).strip().lower()
                    # Normalize header names
                    if 'name' in header_name:
                        header_columns[col_idx] = 'name'
                    elif 'country' in header_name:
                        header_columns[col_idx] = 'country'
                    elif 'postal' in header_name or 'code' in header_name:
                        header_columns[col_idx] = 'postal_code'
                    elif 'exclude' in header_name:
                        header_columns[col_idx] = 'exclude'
                    else:
                        header_columns[col_idx] = header_name
            
            print(f"   [DEBUG]     Mapped header columns: {header_columns}")
            continue
        
        # If we're in a section and have header columns, parse the data row
        if current_section and header_columns:
            rule_data = {
                'section': current_section,
                'name': None,
                'country': None,
                'postal_code': None,
                'exclude': None
            }
            
            # Extract values based on header columns
            for col_idx, header_name in header_columns.items():
                if col_idx < len(row_values):
                    value = row_values[col_idx]
                    if value is not None:
                        rule_data[header_name] = str(value).strip() if value else None
            
            # Only add if we have at least a name or postal code
            if rule_data['name'] or rule_data['postal_code'] or rule_data['country']:
                result['raw_rules'].append(rule_data)
                print(f"   [DEBUG]     + Rule added: {rule_data}")
                
                # Add to appropriate section list
                if current_section == 'postal_code_zones':
                    result['postal_code_zones'].append(rule_data)
                elif current_section == 'country_regions':
                    result['country_regions'].append(rule_data)
                elif current_section == 'no_data_added':
                    result['no_data_added'].append(rule_data)
    
    print(f"\n{'='*60}")
    print(f"[DEBUG] BUSINESS RULES SUMMARY")
    print(f"{'='*60}")
    print(f"   - Postal Code Zones: {len(result['postal_code_zones'])} rules")
    print(f"   - Country Regions: {len(result['country_regions'])} rules")
    print(f"   - No Data Added: {len(result['no_data_added'])} entries")
    print(f"   - Total raw rules: {len(result['raw_rules'])}")
    
    if not result['raw_rules']:
        print(f"\n   [WARNING] No rules were found! Possible issues:")
        print(f"      1. Markers not found in expected format")
        print(f"      2. Headers not in row above markers")
        print(f"      3. Data structure different than expected")
    
    return result


def transform_business_rules_to_conditions(business_rules):
    """
    Transform parsed business rules into condition format.
    
    Args:
        business_rules (dict): Output from process_business_rules()
    
    Returns:
        dict: Dictionary mapping zone/region names to their conditions
              Format: {zone_name: {'country': 'XX', 'postal_codes': ['12', '34'], 'exclude': bool}}
    """
    conditions = {}
    
    for rule in business_rules.get('raw_rules', []):
        name = rule.get('name')
        if not name:
            continue
        
        section = rule.get('section', '')
        
        # Parse postal codes (comma-separated, possibly with spaces)
        # For country_regions, we don't use postal codes - only country matters
        postal_code_str = rule.get('postal_code', '')
        postal_codes = []
        
        if section != 'country_regions' and postal_code_str:
            # Split by comma and clean up each code
            postal_codes = [code.strip() for code in str(postal_code_str).split(',') if code.strip()]
        
        # Excluded column: keep raw value (e.g. "Jiangsu, Jiashan, Jiaxing, Jiangxi") for the result file
        exclude_raw = rule.get('exclude')
        excluded_value = (exclude_raw if exclude_raw is not None else '')
        if not isinstance(excluded_value, str):
            excluded_value = str(excluded_value)
        excluded_value = excluded_value.strip()
        # Boolean for "(EXCLUDE)" in formatted condition: any non-empty value
        is_exclude = bool(excluded_value)

        condition = {
            'section': section,
            'country': rule.get('country'),
            'postal_codes': postal_codes,
            'exclude': is_exclude,
            'excluded_value': excluded_value,  # raw text from Excluded column for result
            'raw_postal_code': postal_code_str if section != 'country_regions' else ''
        }
        
        conditions[name] = condition
    
    return conditions


def format_business_rule_condition(rule_name, condition):
    """
    Format a business rule condition into a readable string.
    
    Args:
        rule_name (str): Name of the rule/zone
        condition (dict): Condition dictionary from transform_business_rules_to_conditions
    
    Returns:
        str: Human-readable condition string
    """
    parts = []
    
    if condition.get('country'):
        parts.append(f"Country: {condition['country']}")
    
    if condition.get('postal_codes'):
        prefix_list = ', '.join(condition['postal_codes'][:5])
        if len(condition['postal_codes']) > 5:
            prefix_list += f", ... (+{len(condition['postal_codes']) - 5} more)"
        parts.append(f"Postal codes starting with: {prefix_list}")
    
    if condition.get('exclude'):
        parts.append("(EXCLUDE)")
    
    return ' | '.join(parts) if parts else 'No conditions'


def find_business_rule_columns(rate_card_df, business_rules_conditions):
    """
    Find which columns in the rate card contain business rule values.
    
    Args:
        rate_card_df (pd.DataFrame): The rate card dataframe
        business_rules_conditions (dict): Dictionary of business rule conditions with rule names as keys
    
    Returns:
        dict: Dictionary with:
            - 'rule_to_columns': {rule_name: [list of columns where found]}
            - 'column_to_rules': {column_name: [list of rules found in it]}
            - 'unique_columns': set of unique column names that contain any business rule
    """
    rule_names = list(business_rules_conditions.keys())
    
    result = {
        'rule_to_columns': {},  # Which columns contain each rule
        'column_to_rules': {},  # Which rules are in each column
        'unique_columns': set()
    }
    
    if rate_card_df is None or rate_card_df.empty or not rule_names:
        return result
    
    print(f"\n{'='*60}")
    print(f"[DEBUG] FINDING BUSINESS RULE COLUMNS IN RATE CARD")
    print(f"{'='*60}")
    print(f"   Searching for {len(rule_names)} rule names in {len(rate_card_df.columns)} columns...")
    
    # Columns to EXCLUDE from business rule detection (these contain codes, not business rule names)
    # Even if a value matches a rule name, these columns should not be treated as business rule columns
    EXCLUDED_BUSINESS_RULE_COLUMNS = {
        'origin airport', 'destination airport', 'origin port', 'destination port',
        'pol', 'poe', 'port of loading', 'port of entry', 'airport', 'port',
        'origin airport code', 'destination airport code', 'airport code',
        'origin seaport', 'destination seaport', 'seaport',
        'ship_port', 'cust_port', 'origin_airport', 'destination_airport',
        'carrier', 'carrier name', 'carrier code', 'scac', 'scac code',
        'origin country', 'destination country', 'country', 'ship_country', 'cust_country'
    }
    
    # Create a set of rule names for faster lookup (case-insensitive)
    rule_names_lower = {str(name).lower(): name for name in rule_names}
    
    # For each column, check which rule names are present
    for col in rate_card_df.columns:
        # Skip excluded columns
        col_lower = str(col).lower().strip()
        if col_lower in EXCLUDED_BUSINESS_RULE_COLUMNS:
            print(f"   [SKIP] Column '{col}' excluded from business rule detection")
            continue
        try:
            # Get unique values in this column
            unique_values = rate_card_df[col].dropna().unique()
            
            # Check each unique value against rule names
            for val in unique_values:
                val_str = str(val).strip().lower()
                
                if val_str in rule_names_lower:
                    original_rule_name = rule_names_lower[val_str]
                    
                    # Track rule to columns mapping
                    if original_rule_name not in result['rule_to_columns']:
                        result['rule_to_columns'][original_rule_name] = []
                    if col not in result['rule_to_columns'][original_rule_name]:
                        result['rule_to_columns'][original_rule_name].append(col)
                    
                    # Track column to rules mapping
                    if col not in result['column_to_rules']:
                        result['column_to_rules'][col] = []
                    if original_rule_name not in result['column_to_rules'][col]:
                        result['column_to_rules'][col].append(original_rule_name)
                    
                    result['unique_columns'].add(col)
        except Exception as e:
            # Skip columns that can't be processed
            pass
    
    # Initialize empty lists for rules not found
    for rule_name in rule_names:
        if rule_name not in result['rule_to_columns']:
            result['rule_to_columns'][rule_name] = []
    
    # Print results
    print(f"\n   [RESULT] Unique columns containing business rules:")
    if result['unique_columns']:
        for col in sorted(result['unique_columns']):
            rules_in_col = result['column_to_rules'].get(col, [])
            print(f"      - '{col}': {len(rules_in_col)} rules found")
            # Show first few rules as examples
            if rules_in_col:
                examples = rules_in_col[:3]
                if len(rules_in_col) > 3:
                    print(f"         Examples: {examples} ... (+{len(rules_in_col) - 3} more)")
                else:
                    print(f"         Rules: {examples}")
    else:
        print(f"      No columns found containing business rule values")
    
    print(f"\n   [SUMMARY] {len(result['unique_columns'])} unique columns contain business rule values")
    
    return result


def get_business_rules_lookup(file_path):
    """
    Get a lookup dictionary from business rule names to their country and postal codes.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
    
    Returns:
        dict: Dictionary with:
            - 'rule_to_country': {rule_name: country_code}
            - 'rule_to_postal_codes': {rule_name: [list of postal codes]}
            - 'business_rule_columns': set of column names containing business rules
            - 'all_rules': list of all rule data with name, country, postal_codes
    """
    # Process business rules
    business_rules = process_business_rules(file_path)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    # Get rate card to find which columns contain business rules
    rate_card_df, rate_card_columns, _ = process_rate_card(file_path)
    business_rule_cols_info = find_business_rule_columns(rate_card_df, business_rules_conditions)
    
    result = {
        'rule_to_country': {},
        'rule_to_postal_codes': {},
        'business_rule_columns': business_rule_cols_info.get('unique_columns', set()),
        'column_to_rules': business_rule_cols_info.get('column_to_rules', {}),
        'all_rules': []
    }
    
    for rule_name, condition in business_rules_conditions.items():
        country = condition.get('country')
        postal_codes = condition.get('postal_codes', [])
        
        if country:
            result['rule_to_country'][rule_name] = country
        if postal_codes:
            result['rule_to_postal_codes'][rule_name] = postal_codes
        
        result['all_rules'].append({
            'name': rule_name,
            'country': country,
            'postal_codes': postal_codes,
            'section': condition.get('section'),
            'exclude': condition.get('exclude', False)
        })
    
    print(f"\n[DEBUG] Business Rules Lookup:")
    print(f"   - Rules with country: {len(result['rule_to_country'])}")
    print(f"   - Rules with postal codes: {len(result['rule_to_postal_codes'])}")
    print(f"   - Columns containing rules: {sorted(result['business_rule_columns'])}")
    
    return result


def get_required_geo_columns():
    """
    Get the list of required geographic columns that should be in the final output.
    These are derived from business rules and should be mapped from ETOF/LC files.
    
    Returns:
        list: List of required column names for origin/destination country and postal codes
    """
    return [
        'Origin Country',
        'Origin Postal Code', 
        'Destination Country',
        'Destination Postal Code'
    ]


def clean_condition_text(condition_text):
    """
    Clean up condition text for better readability.
    
    Transforms:
        "Conditional rules:
        1. 33321-6422: TOPOSTALCODE starts with 33321-6422,333216422"
    To:
        "1. 33321-6422: starts with 33321-6422,333216422"
    """
    import re
    
    if not condition_text:
        return condition_text
    
    # Remove "Conditional rules:" header (case insensitive)
    cleaned = re.sub(r'(?i)^conditional\s*rules\s*:\s*\n?', '', condition_text.strip())
    
    # Remove column name references like "TOPOSTALCODE ", "FROMPOSTALCODE ", etc.
    # Pattern: After the colon and value identifier, remove uppercase column names followed by space
    # Example: "33321-6422: TOPOSTALCODE starts with" -> "33321-6422: starts with"
    cleaned = re.sub(r':\s*[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r': \1', cleaned)
    
    # Also handle cases without numbered format
    cleaned = re.sub(r'^[A-Z_]+\s+(starts with|contains|equals|is empty|does not contain|does not equal)', r'\1', cleaned, flags=re.MULTILINE)
    
    # Clean up extra whitespace and newlines
    lines = [line.strip() for line in cleaned.split('\n') if line.strip()]
    cleaned = '\n'.join(lines)
    
    return cleaned


def _sanitize_for_json(obj):
    """Convert NaN/NaT and non-JSON-serializable values for json.dump."""
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_for_json(x) for x in obj]
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    if pd.isna(obj):
        return None
    if hasattr(obj, 'isoformat'):  # datetime, date, time
        return obj.isoformat()
    return obj


def _has_business_rule_for_cell(rule_name_value, column_name, business_rules_data):
    """
    True if there exists a business rule where Rule Name equals rule_name_value
    and Rate Card Columns contains column_name.
    """
    if rule_name_value is None or (isinstance(rule_name_value, str) and not rule_name_value.strip()):
        return False
    val_str = str(rule_name_value).strip()
    col_str = str(column_name).strip()
    for rule in business_rules_data:
        r_name = rule.get('Rule Name')
        r_cols = rule.get('Rate Card Columns', '')
        if r_name is None or pd.isna(r_name):
            continue
        if str(r_name).strip() != val_str:
            continue
        # Rate Card Columns can be "Origin City" or "Destination City, Origin City"
        cols = [c.strip() for c in str(r_cols).split(',') if c and c.strip()]
        if col_str in cols:
            return True
    return False


def _has_conditional_rule_for_cell(cell_value, column_name, conditions_data):
    """
    True if in conditions_data the row for this column has a Condition Rule
    that contains the cell value (substring, case-insensitive).
    """
    if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
        return False
    val_str = str(cell_value).strip().lower()
    col_str = str(column_name).strip()
    for cond in conditions_data:
        c_col = cond.get('Column')
        if c_col is None or str(c_col).strip() != col_str:
            continue
        rule_text = cond.get('Condition Rule') or ''
        if val_str in str(rule_text).lower():
            return True
    return False


def _enrich_rate_card_records_for_json(rate_card_records, column_names, conditions_data, business_rules_data):
    """
    For each record, for each column except "Lane #", add two keys:
    - "{ColumnName} - Has Business Rule": "Yes"/"No"
    - "{ColumnName} - Has conditional Rule": "Yes"/"No"
    """
    EXCLUDED_COLUMN = "Lane #"
    enriched = []
    for record in rate_card_records:
        new_record = []
        for col in column_names:
            new_record.append((col, record.get(col)))
            if col == EXCLUDED_COLUMN:
                continue
            cell_value = record.get(col)
            has_br = _has_business_rule_for_cell(cell_value, col, business_rules_data)
            has_cr = _has_conditional_rule_for_cell(cell_value, col, conditions_data)
            new_record.append((f"{col} - Has Business Rule", "Yes" if has_br else "No"))
            new_record.append((f"{col} - Has conditional Rule", "Yes" if has_cr else "No"))
        enriched.append(dict(new_record))
    return enriched


def save_rate_card_to_json(rate_card_dataframe, conditions_data, business_rules_data, summary_data,
                           output_path=None, folder_name="partly_df"):
    """
    Save rate card output to a JSON file from the same in-memory data (not from the Excel file).

    Args:
        rate_card_dataframe (pd.DataFrame): Rate card data
        conditions_data (list): List of dicts with Column, Has Condition, Condition Rule
        business_rules_data (list): List of dicts for Business Rules sheet
        summary_data (dict or list): Summary metrics
        output_path (str): Optional full path. If None, uses folder_name + default filename
        folder_name (str): Folder under script directory (default: partly_df)

    Returns:
        str: Path to the saved JSON file
    """
    script_dir = os.path.dirname(os.path.abspath(__file__))
    partly_df_folder = os.path.join(script_dir, folder_name)
    if not os.path.exists(partly_df_folder):
        os.makedirs(partly_df_folder)

    if output_path is None:
        output_path = os.path.join(partly_df_folder, "Filtered_Rate_Card_with_Conditions.json")
    elif not output_path.lower().endswith('.json'):
        output_path = os.path.join(os.path.dirname(output_path), os.path.splitext(os.path.basename(output_path))[0] + '.json')

    rate_card_records = rate_card_dataframe.to_dict(orient='records')
    column_names = list(rate_card_dataframe.columns)
    rate_card_records = _enrich_rate_card_records_for_json(
        rate_card_records, column_names, conditions_data, business_rules_data
    )
    if isinstance(summary_data, dict) and 'Metric' in summary_data:
        summary_dict = dict(zip(summary_data['Metric'], summary_data['Value']))
    else:
        summary_dict = summary_data if isinstance(summary_data, dict) else {}
    payload = {
        'rate_card_data': _sanitize_for_json(rate_card_records),
        'conditions': _sanitize_for_json(conditions_data),
        'business_rules': _sanitize_for_json(business_rules_data),
        'summary': _sanitize_for_json(summary_dict)
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(payload, f, indent=2, ensure_ascii=False, default=str)

    print(f"   JSON saved to: {output_path}")
    return output_path


def save_rate_card_output(file_path, output_path=None, save_excel=True, save_json=True):
    """
    Process rate card and save output to Excel and/or JSON.
    
    Args:
        file_path (str): Path to the rate card file relative to "input/" folder
        output_path (str): Optional output path. If None, saves to "Filtered_Rate_Card_with_Conditions.xlsx"
        save_excel (bool): If True, write the Excel result file (default True)
        save_json (bool): If True, write the JSON result file (default True)
    
    Returns:
        str: Path to the saved Excel file (or output_path even if only JSON was written)
    """
    # Process the rate card
    rate_card_dataframe, rate_card_column_names, rate_card_conditions = process_rate_card(file_path)
    
    # Process business rules
    business_rules = process_business_rules(file_path)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    # Find which columns in rate card contain business rule values
    business_rule_columns = find_business_rule_columns(rate_card_dataframe, business_rules_conditions)
    
    # Set output path - save to partly_df folder (relative to script location)
    if output_path is None:
        # Get the directory where this script is located
        script_dir = os.path.dirname(os.path.abspath(__file__))
        # Ensure partly_df folder exists in the script's directory
        partly_df_folder = os.path.join(script_dir, "partly_df")
        if not os.path.exists(partly_df_folder):
            os.makedirs(partly_df_folder)
        output_path = os.path.join(partly_df_folder, "Filtered_Rate_Card_with_Conditions.xlsx")
    
    # Create conditions DataFrame with cleaned condition text
    conditions_data = []
    for col_name in rate_card_column_names:
        raw_condition = rate_card_conditions.get(col_name, "")
        cleaned_condition = clean_condition_text(raw_condition) if raw_condition else ""
        conditions_data.append({
            'Column': col_name,
            'Has Condition': 'Yes' if col_name in rate_card_conditions else 'No',
            'Condition Rule': cleaned_condition
        })
    
    df_conditions = pd.DataFrame(conditions_data)
    
    # Build business_rules_data and summary_data (used for Excel and/or JSON)
    business_rules_data = []
    for rule_name, condition in business_rules_conditions.items():
        rule_columns = business_rule_columns['rule_to_columns'].get(rule_name, [])
        columns_str = ', '.join(rule_columns) if rule_columns else '(not found in data)'
        exclude_display = condition.get('excluded_value', '') or 'No'
        business_rules_data.append({
            'Rule Name': rule_name,
            'Section': condition.get('section', '').replace('_', ' ').title(),
            'Country': condition.get('country', ''),
            'Postal Codes': condition.get('raw_postal_code', ''),
            'Exclude': exclude_display,
            'Rate Card Columns': columns_str,
            'Formatted Condition': format_business_rule_condition(rule_name, condition)
        })
    
    unique_cols_list = sorted(business_rule_columns['unique_columns']) if business_rule_columns['unique_columns'] else ['(none)']
    summary_data = {
        'Metric': [
            'Total Rows',
            'Total Columns',
            'Columns with Conditions',
            'Columns without Conditions',
            'Business Rules - Postal Code Zones',
            'Business Rules - Country Regions',
            'Business Rules - No Data Added',
            'Columns Using Business Rules',
            'Business Rule Column Names',
            'Source File'
        ],
        'Value': [
            len(rate_card_dataframe),
            len(rate_card_column_names),
            len(rate_card_conditions),
            len(rate_card_column_names) - len(rate_card_conditions),
            len(business_rules.get('postal_code_zones', [])),
            len(business_rules.get('country_regions', [])),
            len(business_rules.get('no_data_added', [])),
            len(business_rule_columns['unique_columns']),
            ', '.join(unique_cols_list),
            file_path
        ]
    }
    
    if save_excel:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            rate_card_dataframe.to_excel(writer, sheet_name='Rate Card Data', index=False)
            df_conditions.to_excel(writer, sheet_name='Conditions', index=False)
            df_business_rules = pd.DataFrame(business_rules_data)
            if not df_business_rules.empty:
                df_business_rules.to_excel(writer, sheet_name='Business Rules', index=False)
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary', index=False)
        print(f"\n✅ Excel saved to: {output_path}")
        print(f"   - Sheet 'Rate Card Data': {len(rate_card_dataframe)} rows x {len(rate_card_column_names)} columns")
        print(f"   - Sheet 'Conditions': {len(rate_card_conditions)} columns with conditions")
        print(f"   - Sheet 'Business Rules': {len(business_rules_conditions)} rules")
        print(f"   - Sheet 'Summary': Overview statistics")
    
    if save_json:
        json_path = os.path.join(os.path.dirname(output_path), os.path.splitext(os.path.basename(output_path))[0] + '.json')
        save_rate_card_to_json(rate_card_dataframe, conditions_data, business_rules_data, summary_data, output_path=json_path)
    
    return output_path


if __name__ == "__main__":
    # Set the input file to process (change this to switch files)
    INPUT_FILE = "Rate Card Export - RA20241217021 v.11.xlsx"
    
    # Choose what to update: Excel, JSON, or both
    SAVE_EXCEL = False   # set False to skip writing the Excel result file
    SAVE_JSON = True   # set False to skip writing the JSON result file
    
    # Process and save (Excel and/or JSON according to flags above)
    output_file = save_rate_card_output(INPUT_FILE, save_excel=SAVE_EXCEL, save_json=SAVE_JSON)
    
    # Also print to console
    rate_card_dataframe, rate_card_column_names, rate_card_conditions = process_rate_card(INPUT_FILE)
    print("\nDataFrame shape:", rate_card_dataframe.shape)
    print("\nColumn names:")
    print(rate_card_column_names)
    print("\nConditions (cleaned):")
    for col, condition in rate_card_conditions.items():
        cleaned = clean_condition_text(condition)
        print(f"  {col}: {cleaned[:100]}..." if len(cleaned) > 100 else f"  {col}: {cleaned}")
    
    # Print Business Rules
    print("\n" + "="*60)
    print("BUSINESS RULES")
    print("="*60)
    business_rules = process_business_rules(INPUT_FILE)
    business_rules_conditions = transform_business_rules_to_conditions(business_rules)
    
    print(f"\nParsed {len(business_rules_conditions)} business rules:")
    for rule_name, condition in business_rules_conditions.items():
        formatted = format_business_rule_condition(rule_name, condition)
        print(f"  {rule_name}: {formatted}")
    
    # Find and print which columns contain business rules
    print("\n" + "="*60)
    print("BUSINESS RULE COLUMNS IN RATE CARD")
    print("="*60)
    business_rule_columns = find_business_rule_columns(rate_card_dataframe, business_rules_conditions)
    
    print(f"\nUnique columns containing business rule values:")
    for col in sorted(business_rule_columns['unique_columns']):
        rules_count = len(business_rule_columns['column_to_rules'].get(col, []))
        print(f"  - {col}: {rules_count} rules")
